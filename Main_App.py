import tkinter as tk
from tkinter import filedialog, IntVar, StringVar
import pandas as pd
from sqlalchemy.types import Text, Float, Integer
import numpy as np
from openpyxl import load_workbook
from sqlalchemy.ext.automap import automap_base
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy import create_engine, String, Column, Table, MetaData
from sklearn import preprocessing, neighbors, model_selection
from PIL import ImageTk, Image
import re


class MainApp:
    def file_import(self):  # funkcja importująca plik źródłowy
        self.path = filedialog.askopenfilename(filetypes=(
            ("Excel files", ".xlsx .xls"), ("all_files", "*.*"), ("xlsx", "*.xlsx"), ("xls", "*.xls")))
        self.main_tab = pd.read_excel(self.path, header=None)
        sheets = pd.ExcelFile(self.path).sheet_names
        if len(sheets) > 1:
            root_f = tk.Toplevel()
            root_f.wm_attributes('-topmost', 1)
            lab = tk.Label(root_f, text='Select sheet:')
            lab.pack()
            sheet_n = tk.StringVar()  # inicjalizacja zmiennej TKINTER
            sheet_n.set(sheets[0])
            for text in sheets:
                tk.Radiobutton(root_f, text=text, variable=sheet_n, value=text,
                                   command=lambda: sheet_n.set(sheet_n.get())).pack(anchor="w")
            exit_button = tk.Button(root_f, text='ok', command=root_f.destroy)
            exit_button.pack()
            root_f.wait_window()
            self.sheet_name = sheet_n.get()
        else:
            self.sheet_name = sheets[0]

        self.main_tab = pd.read_excel(self.path, header=None, sheet_name=self.sheet_name)
        self.belka = tk.Label(text=f"uploaded file: {self.path}", anchor="w", bg='gray83')
        self.belka.place(x=250, y=35, height=20, width=480)

    def file_export(self):

        def source_file_processing(main_tab, path):  # obróbka pliku źródłowegoo
            header_row = main_tab[3].first_valid_index()  # znajduję pierwszy nienullowy wiersz
            del main_tab
            main_tab = pd.read_excel(path, header=header_row, sheet_name=self.sheet_name)  # wczytuję ponownie od nienullowego wiersza
            main_tab.dropna(thresh=len(main_tab.columns) - 4, inplace=True)  # usuwam stopkę - częściowo pusty wiersz
            ind = np.where(main_tab.columns.str.contains('godz', regex=True, flags=re.IGNORECASE))  # znalezienie indeksu nagłówka do zmiany
            main_tab.rename(columns={main_tab.columns[ind[0][0]]: 'godzina'}, inplace=True)  # zamiany nazw nagłówków
            ind = np.where(main_tab.columns.str.contains('GRP', regex=True, flags=re.IGNORECASE))
            main_tab.rename(columns={main_tab.columns[ind[0][0]]: 'GRP'}, inplace=True)
            ind = np.where(main_tab.columns.str.contains('^kana', regex=True, flags=re.IGNORECASE))
            main_tab.rename(columns={main_tab.columns[ind[0][0]]: 'Channel'}, inplace=True)
            main_tab.godzina = main_tab.godzina.astype(int)  # Zmiana formatu godziny
            main_tab.drop(main_tab.filter(regex='Unnamed'), axis=1, inplace=True)
            return main_tab

        def get_all_targets(columns):
            all_targets = columns[2:]
            return all_targets

        def target_selecting(columns):  # wybór grupy referencyjnej
            root_b = tk.Toplevel()
            root_b.wm_attributes('-topmost', 1)
            app_target = tk.StringVar()  # inicjalizacja zmiennej TKINTER
            app_target.set(columns[2])  # ustawienie wartości zmiennej
            lab = tk.Label(root_b, text='Select a reference target')
            lab.pack()
            columns[2:]
            for text in columns[2:]:
                tk.Radiobutton(root_b, text=text, variable=app_target, value=text,
                               command=lambda: app_target.set(app_target.get())).pack(anchor="w")
            exit_button = tk.Button(root_b, text='Potwierdzam to', command=root_b.destroy)
            exit_button.pack()
            root_b.wait_window()
            selected_target = app_target.get()
            return selected_target

        def indexing(grp_tab, selected_target):  # tworzenie indeksów grup na podstawie danych z NIELSENA
            temp_ind = grp_tab.groupby(
                ['Channel']).sum()  # tabelka pomocnicza do indeksów dobowych(do wypełniania braków)
            temp_ind.drop(columns='godzina', inplace=True)
            target_list = []
            for i in grp_tab.columns[2:]:
                if i != selected_target:
                    new_col = 'ind_' + i
                    aux_temp_col = 'aux_ind_' + i
                    grp_tab[new_col] = grp_tab[i] / grp_tab[selected_target]
                    temp_ind[aux_temp_col] = temp_ind[i] / temp_ind[selected_target]
                    grp_tab = pd.merge(grp_tab, temp_ind[[aux_temp_col]], right_index=True, on='Channel')
                    grp_tab[new_col].fillna(grp_tab[aux_temp_col], inplace=True)
                    target_list.append(i)
            return grp_tab, target_list

        def slownik_import():  # import słowników
            src_slownik = r"C:\Users\Michał\Documents\tabele\slownik_zw.xlsx"
            slownik_channels = pd.read_excel(src_slownik, sheet_name='channels')
            slownik_grp = pd.read_excel(src_slownik, header=None, sheet_name='Temp_GRP')
            header_row = slownik_grp[slownik_grp.iloc[:, 0] == 'Channel'].index[0]
            df_temp = slownik_grp.iloc[:, 1]
            df_temp = df_temp[df_temp.str.contains("Universe") == True]
            df_temp.reset_index(drop=True, inplace=True)
            targets_tab = pd.DataFrame()
            targets_tab['Target'] = df_temp.str.split(r"\[|Universe: ", expand=True)[0]
            targets_tab['Universe'] = df_temp.str.split("Universe: ", expand=True)[1].str.rsplit("Cases:", expand=True)[
                0]
            # targets_tab['Target'] = df_temp[0].str.rstrip()
            targets_tab['Universe'] = targets_tab['Universe'].str.replace(u"\xa0", "")
            targets_tab['Target'] = targets_tab['Target'].str.replace(u"\xa0", "")
            targets_tab['Target'] = targets_tab['Target'].astype(str)
            targets_tab['Target'] = targets_tab['Target'].str.rstrip()
            targets_tab['Universe'] = targets_tab['Universe'].astype(int)
            targets_tab = targets_tab.transpose()
            targets_tab.columns = targets_tab.iloc[0, :]
            targets_tab.drop(['Target'], inplace=True)
            targets_tab = targets_tab.astype(int)
            del slownik_grp
            slownik_grp = pd.read_excel(src_slownik, header=header_row, sheet_name='Temp_GRP')
            return slownik_channels, slownik_grp, targets_tab

        def slownik_processing(slownik_grp, slownik_channels):  # obróbka słowników
            slownik_grp.rename(columns={'Day part\Target': 'godzina'}, inplace=True)
            slownik_grp['Channel'] = slownik_grp['Channel'].map(slownik_channels.set_index('arianna_name')['sobr_name'])
            # zamiana nazw ariannowych na sobrowe
            left = slownik_grp['godzina'].str[:2]  # wyłuskanie godzin z ariannowej nazwy pasma
            slownik_grp['godzina'] = pd.to_numeric(left)
            selected_target = target_selecting(slownik_grp.columns)
            print('Reference Target is: ' + selected_target)
            slownik_grp, target_list = indexing(slownik_grp, selected_target)
            return slownik_grp, target_list, selected_target

        def data_merging(main_tab, slownik_channels, slownik_grp, target_list, selected_target,
                         targets_tab):
            main_tab['channel_group'] = main_tab['Channel'].map(
                slownik_channels.set_index('sobr_name')['group_name'])  # left join w inny sposób
            main_tab['kod'] = main_tab['godzina'].map(str) + main_tab[
                'Channel']  # tworzenie identyfikatora(klucz główny) do łączenia indeksu
            slownik_grp['kod'] = slownik_grp['godzina'].astype(str) + slownik_grp[
                'Channel']  # identyfikator w drugiej tabeli
            grp_abs_col = 'liczba_kontaktów_' + selected_target
            main_tab[grp_abs_col] = round(main_tab['GRP'] / 100 * targets_tab[selected_target][0], 0).astype(int)
            main_tab[grp_abs_col].fillna(0, inplace=True)
            for i in target_list:
                index_column = 'ind_' + i
                grp_abs_col = 'liczba_kontaktów_' + i
                aux_temp_col = 'aux_ind_' + i
                grp_col = 'GRP_' + i
                main_tab[index_column] = main_tab['kod'].map(
                    slownik_grp.set_index('kod')[index_column])  # kolumna z indeksami poprzez zmapowanie słownika
                aux_index_pd = slownik_grp.set_index('Channel')[
                    aux_temp_col].drop_duplicates()  # tymczasowa tabela do przypisywania indeksó dobowych w celu uzupełnienia braków
                main_tab[index_column] = main_tab[index_column].fillna(
                    main_tab['Channel'].map(aux_index_pd))  # braki zmapowane tabelą pomocniczą
                main_tab[grp_col] = main_tab['GRP'] * main_tab[index_column]  # indeks * GRP grupy referencyjnej
                main_tab.drop(columns=[index_column], inplace=True)
                main_tab[grp_abs_col] = round(main_tab[grp_col] / 100 * targets_tab[i][0], 0)
                main_tab = main_tab.replace(np.inf, 0)
                main_tab[grp_abs_col] = main_tab[grp_abs_col].astype(int)
                main_tab[grp_abs_col].fillna(0, inplace=True)
            main_tab.drop(columns='kod', inplace=True)
            return main_tab

        def output_file_export(tab):
            root_c = tk.Toplevel()
            root_c.wm_attributes('-topmost', 1)
            belka_3 = tk.Label(root_c, text="Name your output file:")
            belka_3.pack(side="left")
            vartext = tk.StringVar()
            input_window = tk.Entry(root_c, textvariable=vartext, width=50)
            input_window.pack(side="left")
            input_window.focus()
            exit_c = tk.Button(root_c, text='OK', command=root_c.destroy)
            exit_c.pack(side="left")
            # root_c.mainloop()
            root_c.wait_window()
            name = vartext.get()
            link_out = r"C:/Users/Michał/Desktop/{}.xlsx".format(name)
            tab.to_excel(link_out, index=False)
            return link_out

        self.main_tab = source_file_processing(self.main_tab, self.path)
        self.slownik_channels, self.slownik_grp, self.targets_tab = slownik_import()
        self.all_targets = get_all_targets(self.slownik_grp.columns)
        self.slownik_grp, self.target_list, self.selected_target = slownik_processing(self.slownik_grp,
                                                                                      self.slownik_channels)

        self.main_tab = data_merging(self.main_tab, self.slownik_channels, self.slownik_grp, self.target_list,
                                     self.selected_target, self.targets_tab)

        self.lok = output_file_export(self.main_tab)
        self.belka2 = tk.Label(text=f"exported file: {self.lok}", anchor="w", bg='gray83')
        self.belka2.place(x=250, y=75, height=20, width=480)

    def update_db(self):
        def csv_import(link):
            tab = pd.read_csv(link, decimal=',', delimiter=';', header=2)
            tab.dropna(axis=1, thresh=3, inplace=True)
            # tab.drop(columns=[tab.columns.values[len(tab.columns)-1]], inplace=True)
            src_slownik = r"C:\Users\Michał\Documents\tabele\slownik_zw.xlsx"
            slownik_channels = pd.read_excel(src_slownik, sheet_name='channels')
            tab['Channel'] = tab['Channel'].map(slownik_channels.set_index('arianna_name')['sobr_name'])
            tab.rename(columns={r'Start Time\Variables': 'time', 'Cumulated Reach%': 'reach_1+', 'Freq. 3+': 'reach_3+'},
                       inplace=True)
            tab['reach_1+'] = tab['reach_1+'].str.replace(" %", "")
            tab['reach_1+'] = tab['reach_1+'].str.replace(",", ".").astype(float)
            tab['reach_3+'] = tab['reach_3+'].str.replace(" %", "")
            tab['reach_3+'] = tab['reach_3+'].str.replace(",", ".").astype(float)
            tab['reach_1+'] = tab['reach_1+'] / 100
            tab['reach_3+'] = tab['reach_3+'] / 100
            tab['hour'] = tab.apply(lambda t: int(t.time[0:2]), axis=1)
            tab = tab.assign(daypart=pd.cut(tab.hour, [0, 17, 22, 29], labels=["off", "prime", "off2"]))
            tab.daypart = tab.daypart.str.replace('off2', 'off')

            aux_campaigns = pd.DataFrame(tab['Producer'].drop_duplicates())
            aux_campaigns['id'] = range(1, len(aux_campaigns) + 1)
            aux_campaigns['id'] = aux_campaigns['id'].astype(str)

            tab['campaign_id'] = tab['Producer'].map(aux_campaigns.set_index('Producer')['id'])
            tab_grp = pd.pivot_table(tab, index=["campaign_id", "Target"], columns=['Channel', 'daypart'], values="GRP",
                                     aggfunc='sum')
            col_names = ['_'.join(tups) for tups in list(tab_grp.columns)]
            tab_grp.columns = pd.Index(col_names)
            tab_days = pd.pivot_table(tab, index=["campaign_id", "Target"], values="Date", aggfunc=pd.Series.nunique)
            tab_r1 = pd.pivot_table(tab, index=["campaign_id", "Target"], columns="daypart", values="reach_1+",
                                    aggfunc='max')
            tab_r1.columns = pd.Index(list(tab_r1.columns))  # KOLUMNY TUTAJ PRZESTAJĄ MIEĆ INDEXY KATEGORYZOWANE
            tab_r1['reach_1+'] = tab_r1.max(axis=1)
            tab_r1 = tab_r1.drop(tab_r1.columns[[0, 1]], axis=1)
            tab_r3 = pd.pivot_table(tab, index=["campaign_id", "Target"], columns="daypart", values="reach_3+",
                                    aggfunc='max')
            tab_r3.columns = pd.Index(list(tab_r3.columns))  # KOLUMNY TUTAJ PRZESTAJĄ MIEĆ INDEXY KATEGORYZOWANE
            tab_r3['reach_3+'] = tab_r3.max(axis=1)
            tab_r3 = tab_r3.drop(tab_r3.columns[[0, 1]], axis=1)
            vectors = pd.concat([tab_grp, tab_days, tab_r1, tab_r3], axis=1)
            vectors.fillna(0, inplace=True)
            vectors = vectors.reset_index()
            vectors.drop(columns='campaign_id', inplace=True)
            # row_names = ['_'.join(tups) for tups in list(vectors.index)]
            # vectors.index = pd.Index(row_names)
            vectors.rename(columns={'Date': 'Days'}, inplace=True)
            row_numbers = len(vectors)
            return vectors, row_numbers

        def connect_to_database():
            engine = create_engine('mysql+mysqlconnector://root:nasa12crew@localhost:3306/DB_Campaigns')
            return engine

        def export_vectors_to_db(vectors, engine):
            col_names = list(vectors.columns)
            types_dict = {new_list: Float for new_list in col_names}
            types_dict['Target'] = Text
            types_dict['Days'] = Integer
            types_dict['reach_1+'] = Float
            types_dict['reach_3+'] = Float
            vectors.to_sql(
                'campaigns',
                engine,
                if_exists='append',
                index=False,
                dtype=types_dict
            )
        link = r"C:\Users\Michał\Documents\tabele\s1.csv"
        self.vectors = csv_import(link)
        self.engine = connect_to_database()
        export_vectors_to_db(self.vectors[0], self.engine)
        self.belka2 = tk.Label(text=f"added: {self.vectors[1]} rows", anchor="w", bg='light yellow')
        self.belka2.place(x=250, y=515, height=20, width=480)
        self.vectors = None

    def get_camp_vectors(self, main_tab):
        new_tab = main_tab.assign(daypart=pd.cut(main_tab.godzina, [0, 17, 22, 29], labels=["off", "prime", "off2"]))
        new_tab.daypart = new_tab.daypart.str.replace('off2', 'off')
        ref_tg_grp = 'GRP_' + self.selected_target
        new_tab.rename(columns={'GRP': ref_tg_grp}, inplace=True)
        endo_vectors = pd.DataFrame()
        for i, tg in enumerate(self.all_targets):
            new_tab = new_tab.assign(target=tg)
            temp_tg_GRP = 'GRP_' + tg
            endo_vectors = endo_vectors.append(
                new_tab.pivot_table(index='target', columns=['Channel', 'daypart'], values=temp_tg_GRP, aggfunc='sum'))
        col_names = ['_'.join(tups) for tups in list(endo_vectors.columns)]
        endo_vectors.columns = pd.Index(col_names)
        ind = np.where(new_tab.columns.str.contains('^Dat|dat'))
        new_tab.rename(columns={new_tab.columns[ind[0][0]]: 'Days'}, inplace=True)
        days_number = pd.pivot_table(new_tab, index='target', values="Days", aggfunc=pd.Series.nunique)
        endo_vectors = endo_vectors.assign(Days=days_number.iloc[0, 0])
        return endo_vectors

    def get_small_vectors(self, main_tab):
        new_tab = main_tab.assign(daypart=pd.cut(main_tab.godzina, [0, 17, 22, 29], labels=["off", "prime", "off2"]))
        new_tab.daypart = new_tab.daypart.str.replace('off2', 'off')
        ref_tg_grp = 'GRP_' + self.selected_target
        new_tab.rename(columns={'GRP': ref_tg_grp}, inplace=True)
        small_vectors = pd.DataFrame()
        for i, tg in enumerate(self.all_targets):
            new_tab = new_tab.assign(target=tg)
            temp_tg_GRP = 'GRP_' + tg
            small_vectors = small_vectors.append(
                new_tab.pivot_table(index='target', columns=['channel_group'], values=temp_tg_GRP, aggfunc='sum'))
        # col_names = ['_'.join(tups) for tups in list(small_vectors.columns)]
        # small_vectors.columns = pd.Index(col_names)
        ind = np.where(new_tab.columns.str.contains('^Dat|dat'))
        new_tab.rename(columns={new_tab.columns[ind[0][0]]: 'Days'}, inplace=True)
        days_number = pd.pivot_table(new_tab, index='target', values="Days", aggfunc=pd.Series.nunique)
        small_vectors = small_vectors.assign(Days=days_number.iloc[0, 0])
        return small_vectors

    def see_available_targets(self):
        base = automap_base()
        engine = create_engine('mysql+mysqlconnector://root:nasa12crew@localhost:3306/DB_Campaigns')
        base.prepare(engine, reflect=True)
        campaigns = base.classes.campaigns
        # session = Session(engine)
        session = sessionmaker(bind=engine)()
        # result = session.query(campaigns).all()
        # result = [r.Days for r in session.query(campaigns).all()]
        meta = MetaData()
        # camps = Table('campaigns', meta, autoload=True, autoload_with=engine)
        query = session.query(campaigns.target)
        col_tg = pd.read_sql(query.statement, query.session.bind)
        params = col_tg['target'].unique()
        return params

    def params_selecting(self, params):  # wybór grupy referencyjnej
        root_d = tk.Toplevel()
        root_d.wm_attributes('-topmost', 1)
        lab = tk.Label(root_d, text='Match targets:')
        lab.grid(row=0, column=0, columnspan=2)
        var_dict = {}
        for i, tg in enumerate(self.all_targets):
            var_dict[i] = StringVar()
            if i >= len(params):
                var_dict[i].set(params[-1])
            elif i < len(params):
                var_dict[i].set(params[i])
            l = tk.Label(root_d, text=tg, textvariable=tg)
            l.grid(row=i + 1, column=0, sticky='W')
            tk.OptionMenu(root_d, var_dict[i], *params).grid(row=i + 1, column=1, sticky='W')
        exit_button = tk.Button(root_d, text='OK', command=root_d.destroy)
        exit_button.grid(row=i + 2, column=0, columnspan=2)
        root_d.wait_window()
        matched_tg = {}
        selected_params = []
        for i, tg in enumerate(self.all_targets):
            selected_params.append(var_dict[i].get())
            matched_tg[tg] = var_dict[i].get()
        print(selected_params)
        print(matched_tg)
        return selected_params, matched_tg

    def get_vectors_from_db(self, selected_params):
        Base = automap_base()
        engine = create_engine('mysql+mysqlconnector://root:nasa12crew@localhost:3306/DB_Campaigns')
        Base.prepare(engine, reflect=True)
        campaigns = Base.classes.campaigns
        session = sessionmaker(bind=engine)()
        query = session.query(campaigns).filter(campaigns.target.in_(selected_params))
        exo_vectors = pd.read_sql(query.statement, query.session.bind)
        return exo_vectors

    def get_vectors_from_db_small(self, selected_params):
        Base = automap_base()
        engine = create_engine('mysql+mysqlconnector://root:nasa12crew@localhost:3306/DB_Campaigns')
        Base.prepare(engine, reflect=True)
        campaigns = Base.classes.campaigns
        session = sessionmaker(bind=engine)()
        query = session.query(campaigns).filter(campaigns.target.in_(selected_params))
        exo_vectors = pd.read_sql(query.statement, query.session.bind)
        new_tab = pd.DataFrame()
        new_tab['id'] = exo_vectors['id']
        new_tab['target'] = exo_vectors['target']
        new_tab['TVP_1+2'] = exo_vectors['Pr1_prime'] + exo_vectors['Pr2_prime'] + exo_vectors['Pr1_off']\
                             + exo_vectors['Pr2_off']
        new_tab['TVP_Tem']= exo_vectors.drop(['Pr1_prime', 'Pr1_off', 'Pr2_prime', 'Pr2_off', 'Days', 'id', 'reach_1+',
                                              'reach_3+'], axis=1).sum(axis=1, skipna=True)
        new_tab['Days'] = exo_vectors['Days']
        new_tab['reach_1+'] = exo_vectors['reach_1+']
        new_tab['reach_3+'] = exo_vectors['reach_3+']
        exo_vectors = new_tab
        return exo_vectors

    def use_knn(self, matched_tg, endo_vectors, exo_vectors):
        r1_dict = {}
        r3_dict = {}
        scaler = preprocessing.MinMaxScaler()
        for endo_tg, exo_tg in matched_tg.items():
            temp_x = exo_vectors.loc[exo_vectors['target'] == exo_tg].fillna(0)
            x = temp_x.drop(columns=['reach_1+', 'reach_3+', 'id', 'target'])
            x = x.loc[(x != 0).any(axis=1)]
            temp_cols = x.columns
            y1 = temp_x['reach_1+']
            y2 = temp_x['reach_3+']
            model_r1 = neighbors.KNeighborsRegressor(4, metric='euclidean')
            model_r3 = neighbors.KNeighborsRegressor(4, metric='euclidean')
            # x = scaler.fit_transform(x)
            x = pd.DataFrame(x, columns=temp_cols)
            model_r1.fit(x, y1)
            model_r3.fit(x, y2)
            observed_vectors = endo_vectors.loc[endo_tg, :]
            observed_vectors = pd.DataFrame(observed_vectors)
            observed_vectors = observed_vectors.transpose()
            # observed_vectors = scaler.fit_transform(observed_vectors)
            observed_vectors = pd.concat([x, observed_vectors])
            observed_vectors = observed_vectors.tail(1).fillna(0)
            r1 = model_r1.predict(observed_vectors)
            r3 = model_r3.predict(observed_vectors)
            r1_dict[endo_tg] = r1
            r3_dict[endo_tg] = r3
        reach1 = pd.DataFrame.from_dict(r1_dict).iloc[0, :]
        reach3 = pd.DataFrame.from_dict(r3_dict).iloc[0, :]
        print(reach1)
        print(reach3)
        grp_s = endo_vectors.drop(columns='Days')
        grp_s = grp_s.sum(axis=1)
        grp_sum = pd.DataFrame(grp_s).iloc[:, 0]
        tg_tab = self.targets_tab.iloc[0, :]
        summary_df = pd.DataFrame({'GRP': grp_sum, 'reach% 1+': reach1, 'reach% 3+': reach3, 'universe': tg_tab})
        summary_df['GRP'] = round(summary_df['GRP'], 2).astype(float)
        summary_df['impacts'] = round(summary_df['GRP'] / 100 * summary_df['universe'], 0).astype(int)
        summary_df['reach 1+'] = round(summary_df['reach% 1+'] * summary_df['universe'], 0).astype(int)
        summary_df['reach 3+'] = round(summary_df['reach% 3+'] * summary_df['universe'], 0).astype(int)
        summary_df['OTS'] = round(summary_df['impacts'] / summary_df['reach 1+'], 0).astype(int)
        summary_df = summary_df[['GRP', 'impacts', 'reach% 1+', 'reach 1+', 'reach% 3+', 'reach 3+', 'OTS', 'universe']]
        # summary_df.style.format({'reach% 1+': '{:.2%}'.format, 'reach% 1+': '{:.2%}'.format})
        print(summary_df)
        return summary_df

    def estimate_reach(self):
        self.endo_vectors = self.get_camp_vectors(self.main_tab)
        self.params = self.see_available_targets()
        self.selected_params, self.matched_tg = self.params_selecting(self.params)
        self.exo_vectors = self.get_vectors_from_db(self.selected_params)
        self.summary_df = self.use_knn(self.matched_tg, self.endo_vectors, self.exo_vectors)
        wkb = load_workbook(self.lok)
        wkb.create_sheet('summary_tab')
        writer = pd.ExcelWriter(self.lok, engine='openpyxl')
        writer.wkb = wkb
        writer.sheets = dict((ws.title, ws) for ws in wkb.worksheets)
        self.summary_df.to_excel(writer, sheet_name='summary_tab')
        wkb.save(self.lok)
        self.belka3 = tk.Label(text=f"summary saved in: {self.lok}", anchor="w", bg='gray83')
        self.belka3.place(x=30, y=150, height=20, width=700)

    def estimate_reach_small(self):
        self.endo_vectors = self.get_small_vectors(self.main_tab)
        self.params = self.see_available_targets()
        self.selected_params, self.matched_tg = self.params_selecting(self.params)
        self.exo_vectors = self.get_vectors_from_db_small(self.selected_params)
        self.summary_df = self.use_knn(self.matched_tg, self.endo_vectors, self.exo_vectors)
        wkb = load_workbook(self.lok)
        wkb.create_sheet('summary_tab')
        writer = pd.ExcelWriter(self.lok, engine='openpyxl')
        writer.wkb = wkb
        writer.sheets = dict((ws.title, ws) for ws in wkb.worksheets)
        self.summary_df.to_excel(writer, sheet_name='summary_tab')
        wkb.save(self.lok)
        self.belka4 = tk.Label(text=f"summary saved in: {self.lok}", anchor="w", bg='gray83')
        self.belka4.place(x=30, y=150, height=20, width=700)

    def check_db(self):
        self.params = self.see_available_targets()
        root_e = tk.Toplevel()
        root_e.wm_attributes('-topmost', 1)
        lab = tk.Label(root_e, text='Available targets:')
        lab.pack()
        t=tk.Text(root_e)
        for param in self.params:
            t.insert(0.0, param + '\n')
        t.pack()
        root_e.wait_window()

    def __init__(self, parent):
        self.path = None
        self.main_tab = None
        self.vectors = None
        self.engine = None
        self.target_list = None
        self.selected_target = None
        self.ref_tg = None
        self.endo_vectors = None
        self.exo_vectors = None
        self.new_tab = None
        self.reach1 = None
        self.reach3 = None
        self.all_targets = None
        parent.resizable(False, False)
        self.window_app = tk.Frame(parent, height=450*1.318, width=750)
        self.window_app.winfo_toplevel().title("RAGE: Reach And GRP Estimator")
        self.window_app.pack()
        self.image = Image.open("D:/python/Target_Indexing/tlo.jpg")
        self.image = self.image.resize((300, 300), Image.ANTIALIAS)
        self.bgi = ImageTk.PhotoImage(self.image)
        self.back = tk.Label(parent, image=self.bgi).place(x=1, y=40, relwidth=1, relheight=1)
        self.import_button = tk.Button(parent, text="Upload Your Campaign", command=self.file_import,
                                       activeforeground="purple1", activebackground="pale green",
                                       highlightcolor="pale green", bg='CadetBlue2')
        self.import_button.place(x=30, y=30, height=30, width=200)
        self.process_button = tk.Button(parent, text="Estimate Client's Target GRPs", command=self.file_export, bg='CadetBlue2',
                                       activeforeground="purple1", activebackground="pale green")
        self.process_button.place(x=30, y=70, height=30, width=200)

        self.reach_button = tk.Button(parent, text="Estimate Campaign's Reach", bg='SteelBlue1', command=self.estimate_reach)
        self.reach_button.place(x=30, y=110, height=30, width=200)

        self.small_reach_button = tk.Button(parent, text="Estimate Small Campaign's Reach", bg='SteelBlue1',
                                        command=self.estimate_reach_small)
        self.small_reach_button.place(x=250, y=110, height=30, width=200)

        self.update_button = tk.Button(parent, text="Update Campaigns Database", command=self.update_db, bg='sandy brown')
        self.update_button.place(x=30, y=510, height=30, width=200)
        self.check_button = tk.Button(parent, text="See available targets in Database", command=self.check_db,
                                       bg='sandy brown')
        self.check_button.place(x=250, y=510, height=30, width=200)

root = tk.Tk()
app = MainApp(root)
root.mainloop()


# ref_tg = app.selected_target

# app.path
# app.main_tab
#
# print(type(app.path))